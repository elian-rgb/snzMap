"""Load Kiki's workbook as the gold-standard eval set.

The workbook is the human artifact; this turns it into a machine one without touching it.
Nothing here writes to the .xlsx. Kiki's README has an explicit rule — "don't add new
columns ad hoc" — so `venue_id`, `operator_normalized` and `tenure_id` are DERIVED on
every load rather than stored as columns she would have to maintain.

Resolving `venue_name` -> `venue_id` here is not bookkeeping. It is the first live test of
the Phase 0 alias machinery against names a human typed from memory, which is the same
shape of input the extractor will face from a newspaper. A gold row that cannot be resolved
to the spine is a finding about the spine, not a typo to paper over.

    .venv/bin/python -m pipeline.spans.gold
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl

from ..schema import (
    SPINE_TO_TENURE_TYPE,
    WORKBOOK_HUNT_COLUMNS,
    WORKBOOK_TENURE_COLUMNS,
    normalize_operator,
    validate_hunt,
    validate_tenure,
)
from ..spine.build_spine import match_key

WORKBOOK = Path.home() / "Downloads" / "SNZ_contract_tenure_table.xlsx"
OUT = Path(__file__).resolve().parent.parent / "output"
SPINE = OUT / "venues_full.json"


# ── Cell coercion ──────────────────────────────────────────────────────────────

def _cell(v: Any) -> Any:
    """Excel hands back datetimes, floats and stray whitespace. Dates become YYYY-MM-DD."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _read_sheet(ws, columns: list[str]) -> tuple[list[dict[str, Any]], list[str]]:
    header = [_cell(c.value) for c in ws[1]][: len(columns)]
    problems = []
    if header != columns:
        problems.append(
            f"{ws.title}: header drifted from the locked schema.\n"
            f"  expected: {columns}\n  found:    {header}"
        )
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {c: _cell(v) for c, v in zip(columns, values)}
        if all(v is None for v in row.values()):
            continue  # the workbook is pre-padded to 500 blank rows
        row["_excel_row"] = excel_row
        rows.append(row)
    return rows, problems


# ── Spine resolution ───────────────────────────────────────────────────────────

def _load_spine_index() -> dict[str, list[dict[str, Any]]]:
    venues = json.loads(SPINE.read_text())
    index: dict[str, list[dict[str, Any]]] = {}
    for v in venues:
        for key in v.get("match_keys", {}):
            index.setdefault(key, []).append(v)
    return index


def year_of(value: Any) -> int | None:
    try:
        return int(str(value)[:4])
    except (TypeError, ValueError):
        return None


def open_at(venue: dict[str, Any], key: str, year: int) -> bool:
    """Could this venue, under this name, have been operating in `year`?

    Unknown dates count as plausible. A null opened_date means "we don't know", not
    "never opened", and treating it as a disqualifier would silently drop real venues.
    """
    opened, closed = year_of(venue.get("opened_date")), year_of(venue.get("closed_date"))
    if opened is not None and year < opened:
        return False
    if closed is not None and year > closed:
        return False
    window = venue.get("match_keys", {}).get(key) or {}
    ks, ke = year_of(window.get("start_date")), year_of(window.get("end_date"))
    if ks is not None and year < ks:
        return False
    if ke is not None and year > ke:
        return False
    return True


def resolve_venue(
    row: dict[str, Any],
    index: dict[str, list[dict[str, Any]]],
    year: int | None = None,
) -> dict[str, Any]:
    """Name (+ city/state/year) -> spine venue. Returns the match and how confident it is."""
    name = row.get("venue_name")
    if not name:
        return {"venue_id": None, "match": "no_name", "candidates": []}

    key = match_key(name)
    candidates = index.get(key, [])
    if not candidates:
        return {"venue_id": None, "match": "not_in_spine", "candidates": []}

    if len(candidates) > 1:
        # Two Busch Stadiums, two Wrigley Fields. Disambiguate on what the row already
        # carries rather than silently taking the first — the spine's ambiguity report
        # exists precisely so this cannot be swept under the rug.
        state, city = row.get("state"), row.get("city")
        narrowed = [c for c in candidates if not state or c.get("state") == state]
        if len(narrowed) > 1 and city:
            narrowed = [c for c in narrowed if (c.get("city") or "").lower() == city.lower()] or narrowed

        # Geography is not always enough, and this is not an edge case: the three Madison
        # Square Gardens are all in Manhattan, NY, and the two Busch Stadiums are both in
        # St. Louis. What separates them is time — the 1925 Garden closed in 1968, so a
        # 2005 contract cannot be about it. Any row that carries a date gets this axis.
        if len(narrowed) > 1 and year is not None:
            by_time = [c for c in narrowed if open_at(c, key, year)]
            if by_time:
                narrowed = by_time

        if len(narrowed) != 1:
            return {
                "venue_id": None,
                "match": "ambiguous",
                "candidates": [
                    {"venue_id": c["venue_id"], "canonical_name": c["canonical_name"],
                     "city": c.get("city"), "state": c.get("state"),
                     "opened_date": c.get("opened_date"), "closed_date": c.get("closed_date")}
                    for c in candidates
                ],
            }
        candidates = narrowed

    v = candidates[0]
    exact = key == match_key(v["canonical_name"])
    return {
        "venue_id": v["venue_id"],
        "match": "exact" if exact else "alias",
        "matched_name": v["canonical_name"],
        "spine_type": v.get("venue_type"),
        "spine_city": v.get("city"),
        "spine_state": v.get("state"),
        "spine_lat": v.get("lat"),
        "spine_lng": v.get("lng"),
        "candidates": [],
    }


def _cross_check(row: dict[str, Any], res: dict[str, Any]) -> list[str]:
    """Disagreements between what Kiki typed and what the spine holds. Warnings, not errors —
    she may well be right and the spine wrong."""
    warnings = []
    if not res.get("venue_id"):
        return warnings
    if row.get("state") and res.get("spine_state") and row["state"] != res["spine_state"]:
        warnings.append(f"state: workbook={row['state']} spine={res['spine_state']}")
    expected_type = SPINE_TO_TENURE_TYPE.get(res.get("spine_type") or "", "other")
    if row.get("venue_type") and row["venue_type"] != expected_type:
        warnings.append(
            f"venue_type: workbook={row['venue_type']} "
            f"spine={res['spine_type']} (projects to {expected_type})"
        )
    for axis, wb_key, spine_key in (("lat", "lat", "spine_lat"), ("lng", "lng", "spine_lng")):
        a, b = row.get(wb_key), res.get(spine_key)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)) and abs(a - b) > 0.05:
            warnings.append(f"{axis}: workbook={a} spine={b}")
    return warnings


# ── Load ───────────────────────────────────────────────────────────────────────

def _tenure_id(row: dict[str, Any], venue_id: str | None, operator: str | None) -> str:
    seed = f"{venue_id or row.get('venue_name')}|{operator}|{row.get('start_date')}"
    return "gold-" + hashlib.sha256(seed.encode()).hexdigest()[:10]


def load_gold(workbook: Path = WORKBOOK) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    index = _load_spine_index()

    tenure_rows, problems = _read_sheet(wb["tenure_table"], WORKBOOK_TENURE_COLUMNS)
    hunt_rows, hunt_problems = _read_sheet(wb["hunt_log"], WORKBOOK_HUNT_COLUMNS)
    problems += hunt_problems

    gold = []
    for row in tenure_rows:
        excel_row = row.pop("_excel_row")
        op = normalize_operator(row.get("operator"))
        # The row's own start year is the disambiguating axis when city/state is not enough.
        res = resolve_venue(row, index, year=year_of(row.get("start_date")))
        rec = {
            **row,
            "tenure_id": _tenure_id(row, res["venue_id"], op["operator_normalized"]),
            "venue_id": res["venue_id"],
            "operator_normalized": op["operator_normalized"],
            "derived_from": [],
        }
        row_problems = [f"row {excel_row}: {p}" for p in validate_tenure(rec)]
        gold.append({
            "record": rec,
            "excel_row": excel_row,
            "venue_match": res["match"],
            "venue_candidates": res.get("candidates", []),
            "operator_relation": op["relation"],
            "warnings": _cross_check(row, res),
            "problems": row_problems,
            # The workbook's own convention for "I know this but haven't checked it".
            "verified": "verify" not in (row.get("source") or "").lower(),
        })
        problems += row_problems

    hunt = []
    for row in hunt_rows:
        excel_row = row.pop("_excel_row")
        rec = {**row, "search_id": "hunt-" + hashlib.sha256(
            f"{row.get('venue_or_operator')}|{row.get('period_searched')}|"
            f"{row.get('where_searched')}|{row.get('date_of_search')}".encode()
        ).hexdigest()[:10], "query_used": None}
        row_problems = [f"hunt row {excel_row}: {p}" for p in validate_hunt(rec)]
        hunt.append({"record": rec, "excel_row": excel_row, "problems": row_problems})
        problems += row_problems

    return {"gold": gold, "hunt": hunt, "problems": problems}


def main() -> None:
    result = load_gold()
    gold, hunt, problems = result["gold"], result["hunt"], result["problems"]

    OUT.mkdir(exist_ok=True)
    (OUT / "gold_tenure.json").write_text(json.dumps(gold, indent=2))
    (OUT / "gold_hunt_log.json").write_text(json.dumps(hunt, indent=2))

    verified = [g for g in gold if g["verified"]]
    resolved = [g for g in gold if g["record"]["venue_id"]]
    print(f"tenure rows: {len(gold)}  (verified {len(verified)}, unverified {len(gold) - len(verified)})")
    print(f"resolved to spine: {len(resolved)}/{len(gold)}")
    print(f"hunt_log rows: {len(hunt)}")

    for g in gold:
        if not g["record"]["venue_id"]:
            name = g["record"].get("venue_name")
            extra = ""
            if g["venue_match"] == "ambiguous":
                extra = " -> " + ", ".join(
                    f"{c['venue_id']} ({c['city']}, {c['state']})" for c in g["venue_candidates"]
                )
            print(f"  UNRESOLVED row {g['excel_row']}: {name} [{g['venue_match']}]{extra}")
        for w in g["warnings"]:
            print(f"  warn row {g['excel_row']} ({g['record'].get('venue_name')}): {w}")

    if problems:
        print(f"\n{len(problems)} schema problem(s):")
        for p in problems:
            print(f"  {p}")
    else:
        print("\nschema: clean")

    if len(gold) < 10:
        print(
            f"\nGOLD SET NOT READY: {len(gold)} row(s), the plan calls for 10-20.\n"
            "  These rows are Kiki's to write — they are the standard the pipeline is\n"
            "  measured against, so a row invented by the pipeline's author would make the\n"
            "  eval gate measure nothing. pipeline/spans/evaluate.py runs regardless and\n"
            "  will report recall over whatever is present."
        )


if __name__ == "__main__":
    main()
